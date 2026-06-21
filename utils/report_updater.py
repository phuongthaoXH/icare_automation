"""
utils/report_updater.py
Cập nhật kết quả test vào file Excel TestCase_D02_ThuTuc600_iCare.xlsx
(file test case gốc đã tạo trước đó).
"""
import os
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

logger = logging.getLogger(__name__)

# Map mã TC sang sheet và row
# Key: TC ID prefix → sheet name
SHEET_MAP = {
    "TC_D02_PRE": "01_Giao_Dien_D02",
    "TC_D02_UI":  "01_Giao_Dien_D02",
    "TC_D02_UX":  "01_Giao_Dien_D02",
    "TC_D02_HP":  "02_Chuc_Nang_D02",
    "TC_D02_NEG": "02_Chuc_Nang_D02",
    "TC_D02_EDGE":"02_Chuc_Nang_D02",
    "TC_VAL":     "03_Validate_Fields_D02",
    "TC_PQ":      "04_PhanQuyen_BaoMat",
    "TC_SEC":     "04_PhanQuyen_BaoMat",
    "TC_EXP":     "05_Exploratory_D02",
}

COLOR_MAP = {
    "P":  "C6EFCE",  # Green
    "F":  "FFC7CE",  # Red
    "PE": "FFEB9C",  # Yellow
    "":   "FFFFFF",  # White
}


class ExcelReportUpdater:
    """
    Cập nhật cột Kết quả và Ghi chú trong file Excel test case.
    """

    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self._wb = None

    def _load(self):
        if not os.path.exists(self.excel_path):
            logger.warning(f"File Excel không tồn tại: {self.excel_path}")
            return False
        self._wb = load_workbook(self.excel_path)
        return True

    def update_result(self, tc_id: str, status: str, error_msg: str = ""):
        """
        Cập nhật kết quả cho một test case.

        Args:
            tc_id:     Mã TC, VD 'TC_D02_HP_01'
            status:    'P' | 'F' | 'PE'
            error_msg: Mô tả lỗi nếu F
        """
        if not self._wb and not self._load():
            return

        # Tìm sheet phù hợp
        sheet_name = None
        for prefix, sname in SHEET_MAP.items():
            if tc_id.upper().startswith(prefix):
                sheet_name = sname
                break

        if not sheet_name or sheet_name not in self._wb.sheetnames:
            logger.debug(f"Không tìm được sheet cho TC: {tc_id}")
            return

        ws = self._wb[sheet_name]
        color = COLOR_MAP.get(status, "FFFFFF")
        fill  = PatternFill("solid", fgColor=color)

        for row in ws.iter_rows(min_row=11):
            cell_a = row[0]  # Cột A = Trường hợp
            if cell_a.value and str(cell_a.value).strip().upper() == tc_id.upper():
                # Cột E = KQ Lần 1 (index 4)
                kq_cell = row[4]
                kq_cell.value = status
                kq_cell.fill  = fill
                kq_cell.alignment = Alignment(horizontal="center", vertical="center")
                kq_cell.font  = Font(bold=True)

                # Cột H = KQ hiện tại (index 7)
                row[7].value = status
                row[7].fill  = fill
                row[7].alignment = Alignment(horizontal="center")

                # Cột J = Ghi chú (index 9)
                if error_msg:
                    existing = row[9].value or ""
                    ts = datetime.now().strftime("%d/%m %H:%M")
                    row[9].value = f"[{ts}] {error_msg[:200]}"

                logger.info(f"Updated {tc_id} → {status}")
                return

        logger.debug(f"Không tìm thấy TC '{tc_id}' trong sheet '{sheet_name}'")

    def bulk_update(self, results: list[dict]):
        """
        Cập nhật nhiều kết quả cùng lúc.

        Args:
            results: list of {"tc_id": str, "status": str, "error": str}
        """
        if not self._load():
            return
        for r in results:
            self.update_result(
                tc_id=r.get("tc_id", ""),
                status=r.get("status", "PE"),
                error_msg=r.get("error", ""),
            )
        self._save()

    def _save(self):
        if self._wb:
            self._wb.save(self.excel_path)
            logger.info(f"Saved updated results: {self.excel_path}")

    def __enter__(self):
        self._load()
        return self

    def __exit__(self, *args):
        self._save()
