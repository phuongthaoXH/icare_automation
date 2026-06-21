"""
ExcelReporter – Ghi kết quả test tự động vào file Excel test case.
Cập nhật cột E (KQ Lần 1) trong file TestCase_D02_ThuTuc600_iCare.xlsx
"""
import os
import logging
from datetime import datetime
from typing import Optional

logger = logging.getLogger(__name__)

# Mapping TC_ID → (sheet_name, row_number)
# Sẽ được build tự động khi load file
_TC_MAP: dict[str, tuple[str, int]] = {}


def load_testcase_file(path: str) -> Optional[object]:
    """Load workbook và build map TC_ID → (sheet, row)."""
    global _TC_MAP
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path)
        _TC_MAP.clear()
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=11, values_only=False):
                tc_cell = row[0]  # Column A
                if tc_cell.value and str(tc_cell.value).startswith("TC_"):
                    _TC_MAP[str(tc_cell.value)] = (sheet_name, tc_cell.row)
        logger.info(f"Loaded {len(_TC_MAP)} test cases từ {path}")
        return wb
    except Exception as e:
        logger.warning(f"Không load được Excel: {e}")
        return None


def write_results(path: str, results: list[dict]):
    """
    Ghi kết quả vào cột KQ Lần 1 (col E = 5).
    results: [{tc_id, status, error, duration}]
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment

        if not os.path.exists(path):
            logger.warning(f"File không tồn tại: {path}")
            return

        wb = openpyxl.load_workbook(path)

        color_map = {
            "P":  ("C6EFCE", "006100"),   # xanh lá
            "F":  ("FFC7CE", "9C0006"),   # đỏ
            "PE": ("FFEB9C", "9C6500"),   # vàng
        }

        for r in results:
            tc_id  = r.get("tc_id", "")
            status = r.get("status", "PE")
            error  = r.get("error", "")[:300]

            # Tìm trong _TC_MAP (exact) hoặc partial match
            match = None
            for key in _TC_MAP:
                if key in tc_id or tc_id in key:
                    match = _TC_MAP[key]
                    break

            if not match:
                continue

            sheet_name, row_num = match
            ws = wb[sheet_name]

            bg, fg = color_map.get(status, ("FFFFFF", "000000"))

            # Col E = KQ Lần 1
            cell_e = ws.cell(row=row_num, column=5, value=status)
            cell_e.fill = PatternFill("solid", fgColor=bg)
            cell_e.font = Font(bold=True, color=fg)
            cell_e.alignment = Alignment(horizontal="center")

            # Col H = KQ Hiện tại
            cell_h = ws.cell(row=row_num, column=8, value=status)
            cell_h.fill = PatternFill("solid", fgColor=bg)
            cell_h.font = Font(bold=True, color=fg)
            cell_h.alignment = Alignment(horizontal="center")

            # Col I = Mã lỗi (ngắn)
            if status == "F" and error:
                ws.cell(row=row_num, column=9, value=error[:80])

            # Col J = Thêm thời gian chạy
            duration = r.get("duration", 0)
            ws.cell(row=row_num, column=10,
                    value=f"{r.get('time','')} ({duration:.1f}s)")

        wb.save(path)
        logger.info(f"Đã ghi {len(results)} kết quả vào {path}")

    except Exception as e:
        logger.error(f"Lỗi ghi Excel: {e}")


def generate_summary_sheet(path: str, results: list[dict]):
    """Thêm sheet Summary tổng hợp kết quả."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        wb = openpyxl.load_workbook(path)

        # Xóa sheet cũ nếu có
        if "Summary" in wb.sheetnames:
            del wb["Summary"]

        ws = wb.create_sheet("Summary", 0)

        # Header
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = f"KẾT QUẢ KIỂM THỬ TỰ ĐỘNG – Tờ khai D02 – {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="1F4E79")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Stats
        total = len(results)
        passed = sum(1 for r in results if r["status"] == "P")
        failed = sum(1 for r in results if r["status"] == "F")
        pending = sum(1 for r in results if r["status"] == "PE")
        pass_rate = (passed / total * 100) if total > 0 else 0

        stats = [
            ("Tổng số test case", total,      "4472C4"),
            ("Đạt (P)",           passed,     "70AD47"),
            ("Không đạt (F)",     failed,     "FF0000"),
            ("Pending (PE)",      pending,    "FFC000"),
            ("Tỷ lệ pass",        f"{pass_rate:.1f}%", "4472C4"),
        ]

        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for i, (label, value, color) in enumerate(stats, 3):
            lc = ws.cell(row=i, column=1, value=label)
            vc = ws.cell(row=i, column=2, value=value)
            lc.font = Font(bold=True, color="FFFFFF")
            vc.font = Font(bold=True, color="FFFFFF")
            lc.fill = PatternFill("solid", fgColor=color)
            vc.fill = PatternFill("solid", fgColor=color)
            lc.border = border
            vc.border = border
            lc.alignment = Alignment(horizontal="left", vertical="center")
            vc.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[i].height = 22

        # Detail table
        detail_headers = ["TC ID", "Kết quả", "Thời gian (s)", "Lỗi", "Thực hiện lúc"]
        for ci, h in enumerate(detail_headers, 1):
            hc = ws.cell(row=9, column=ci, value=h)
            hc.font = Font(bold=True, color="FFFFFF")
            hc.fill = PatternFill("solid", fgColor="2E75B6")
            hc.border = border
            hc.alignment = Alignment(horizontal="center")

        color_map = {"P": "C6EFCE", "F": "FFC7CE", "PE": "FFEB9C"}
        for ri, r in enumerate(results, 10):
            row_data = [
                r.get("tc_id", ""), r.get("status", ""),
                round(r.get("duration", 0), 2),
                r.get("error", "")[:100], r.get("time", ""),
            ]
            bg = color_map.get(r.get("status", ""), "FFFFFF")
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = PatternFill("solid", fgColor=bg)
                c.border = border
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Column widths
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 60
        ws.column_dimensions["E"].width = 20

        wb.save(path)
        logger.info(f"Summary sheet added to {path}")

    except Exception as e:
        logger.error(f"Lỗi tạo Summary sheet: {e}")
