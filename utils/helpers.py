"""utils/helpers.py – Tiện ích dùng chung."""
import os, re, time, logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)


def today_vn() -> str:
    return datetime.today().strftime("%d/%m/%Y")

def thang_nam_now() -> str:
    return datetime.today().strftime("%m/%Y")

def add_months(base: str, n: int) -> str:
    month, year = map(int, base.split("/"))
    total = month + n
    year  += (total - 1) // 12
    month  = ((total - 1) % 12) + 1
    return f"{month:02d}/{year}"


def make_d02_excel(filepath: str, rows: int = 3, include_invalid: bool = False) -> str:
    """Tạo file Excel mẫu cho import D02."""
    os.makedirs(os.path.dirname(os.path.abspath(filepath)) or ".", exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "D02"
    headers = [
        "Họ và tên","Loại phương án","Phương án điều chỉnh",
        "Loại ngày sinh","Ngày sinh","Giới tính","CMND/CCCD",
        "Mã số BHXH","Chức vụ","Phòng ban",
        "Số HĐLĐ/Quyết định","Ngày ký",
        "Từ tháng/năm","Đến tháng/năm",
        "Nơi làm việc","Tỷ lệ đóng (%)","Tiền lương","Ghi chú",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center")

    genders = ["Nam","Nữ"]
    for i in range(1, rows + 1):
        ws.append([
            f"Nhân Viên Auto {i:03d}",
            "Tham gia BHXH, BHYT, BHTN","Tăng mới",
            "Ngày/Tháng/Năm",
            f"{(i%28+1):02d}/0{(i%9+1)}/199{i%9+1}",
            genders[i%2],
            f"{i:012d}",
            "","Nhân viên","Phòng Kinh doanh",
            f"HD-AUTO-{i:04d}","01/01/2025",
            "01/2025","12/2025",
            "Hà Nội","8","5000000",
            f"Auto test row {i}",
        ])

    if include_invalid:
        ws.append(["","INVALID","","","99/99/9999","",
                   "123"] + [""]*11)

    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(w+4, 40)

    wb.save(filepath)
    logger.info(f"Excel sample: {filepath} ({rows} rows)")
    return filepath


def retry(fn, times: int = 3, delay: float = 1.0, exceptions=(Exception,)):
    last_exc = None
    for attempt in range(times):
        try:
            return fn()
        except exceptions as e:
            last_exc = e
            if attempt < times - 1:
                time.sleep(delay)
    raise last_exc


class ExcelReporter:
    STATUS_COLOR = {"P":"C6EFCE","F":"FFC7CE","PE":"FFEB9C","":"FFFFFF"}

    def __init__(self, output_path: str = "reports/test_results.xlsx"):
        self.output_path = output_path
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        self.wb = openpyxl.Workbook()
        ws = self.wb.active
        ws.title = "Kết quả test"
        for i, h in enumerate(["TC ID","Mục đích","KQ Lần 1","KQ Hiện tại",
                                "Thời gian (s)","Lỗi","Thực hiện lúc"],1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True, color="FFFFFF", name="Arial")
            c.fill = PatternFill("solid", fgColor="1F4E79")
            c.alignment = Alignment(horizontal="center")

    def write_result(self, tc_id, purpose, status, duration=0, error_msg=""):
        ws = self.wb.active
        ws.append([tc_id, purpose, status, status,
                   round(duration,2),
                   (error_msg or "")[:300],
                   datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
        color = self.STATUS_COLOR.get(status,"FFFFFF")
        for col in range(1, 8):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill = PatternFill("solid", fgColor=color)
            cell.font = Font(name="Arial", size=10)

    def save(self):
        for col in self.wb.active.columns:
            w = max(len(str(c.value or "")) for c in col)
            self.wb.active.column_dimensions[col[0].column_letter].width = min(w+4,60)
        self.wb.save(self.output_path)
        logger.info(f"Saved: {self.output_path}")
        return self.output_path


def normalize_vn(text: str) -> str:
    return re.sub(r"\s+", " ", text.strip())

def mask_sensitive(text: str) -> str:
    return re.sub(r"(?i)(password|token|secret)\s*=\s*\S+", r"\1=***", text)
